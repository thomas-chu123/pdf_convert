import os
import re
import PyPDF2
import openpyxl

title_list = ['Title', 'Case ID', 'Test type', 'Test case coverage', 'Preconditions', 'Test Steps', 'Expected Results']

# Change directory to the folder containing PDF files
# os.chdir('.')
file_data = {}
# Loop through all PDF files in the folder
for filename in os.listdir('.'):
    if filename.endswith('.pdf'):
        # Open the PDF file
        data_list = []
        with open(filename, 'rb') as file:
            # Create a PDF reader object
            reader = PyPDF2.PdfReader(file)
            # Loop through each page of the PDF
            for page in range(len(reader.pages)):
                # Extract the text from the page
                text = reader.pages[page].extract_text()
                # Print the text
                # print(text)
                text_line = text.split('\n')
                for line in text_line:
                    data_list.append(line)
        file_data[filename] = data_list

# print(file_data)

test_case_block = []
test_case = {}
start_case = bool
start_test_type = bool
start_test_coverage = bool
start_preconditions = bool
start_expectations = bool
start_test_steps = bool

workbook = openpyxl.Workbook()

for filename in file_data:
    test_case_block = []
    worksheet = workbook.create_sheet(filename)

    col=1
    for title in title_list:
        worksheet.cell(row=1, column=col, value=title)
        col = col + 1

    for line in file_data[filename]:
        if re.match(r'[1-9]\.\d', line) and "....." not in line:
            test_case = {}
            if test_case != None:
                test_case_block.append(test_case)
            start_expectations = False
            test_case['Title'] = line
        if "Case ID" in line:
            start_case = True
            continue
        if "Test type" in line:
            start_case = False
            start_test_type = True
            continue
        if "Test case coverage" in line:
            start_test_type = False
            start_test_coverage = True
            continue
        if "Preconditions" in line:
            start_test_coverage = False
            start_preconditions = True
            continue
        if "Test Steps" in line:
            start_preconditions = False
            start_test_steps = True
            continue
        if "Expected Results" in line:
            start_test_steps = False
            start_expectations = True
            continue

        if start_case == True:
            test_case['Case ID'] = line
        if start_test_type == True:
            test_case['Test type'] = line
        if start_test_coverage == True:
            test_case['Test case coverage'] = line
        if start_preconditions == True:
            if 'Preconditions' in test_case:
                test_case['Preconditions'] = test_case['Preconditions'] + "\n" + line
            else:
                test_case['Preconditions'] = line
        if start_test_steps == True:
            if 'Test Steps' in test_case:
                test_case['Test Steps'] = test_case['Test Steps'] + "\n" + line
            else:
                test_case['Test Steps'] = line
        if start_expectations == True:
            if 'Expected Results' in test_case:
                test_case['Expected Results'] = test_case['Expected Results'] + "\n" + line
            else:
                test_case['Expected Results'] = line
        # previous_line = line

    row = 2
    col = 1
    for test_case in test_case_block:
        col = 1
        for title in title_list:
            if title in test_case:
                worksheet.cell(row=row, column=col, value=test_case[title])
            else:
                worksheet.cell(row=row, column=col, value="")
            col += 1
        row += 1
workbook.save('testcase.xlsx')
print (test_case_block)

